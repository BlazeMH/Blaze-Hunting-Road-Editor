# core/excel.py

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .constants import MONSTERS, DETAILS_XLSX_DEFAULT


# ----------------------------
# Helpers / Styling
# ----------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")     # deep slate
GROUP_FILL  = PatternFill("solid", fgColor="2E3B5B")     # slightly lighter
ALT_FILL    = PatternFill("solid", fgColor="0D1117")     # very dark (optional)
WHITE_FONT  = Font(color="FFFFFF", bold=True)
LEFT        = Alignment(horizontal="left", vertical="center")
CENTER      = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)

# Typical notes per group index (0-based).
GROUP_HINTS = {
    0:  "All types of Monsters",
    1:  "Usually Burst/Origin Monsters",
    2:  "All types of Monsters",
    3:  "Usually Exotic Monsters",
    4:  "All types of Monsters",
    5:  "Fatalis Floor 1",
    6:  "Exotics/ElderDragons/Origin Mix",
    7:  "Usually Zenith 1 Monsters",
    8:  "Usually Zenith 2 Monsters",
    9:  "Usually Zenith 3 Monsters",
    10: "Usually Zenith 4 Monsters",
    11: "Fatalis Floor 2",
    12: "All Types of Monsters",
    13: "Usually Burst/Origin Monsters",
    14: "Usually Exotic Monsters",
    15: "Usually Zenith 1 Monsters",
    16: "Usually Zenith 2 Monsters",
    17: "Usually Zenith 3 Monsters",
    18: "Usually Zenith 4 Monsters",
    19: "Usually Basic Monsters",
    20: "Usually Mid-level Monsters",
    21: "Usually Mid Level Monsters",
    22: "Usually Elder Dragons"
}

# Regex that detects a group header regardless of extra labels.
# Matches: "-- Group 13 --" OR "-- Group 13 -- Usually Something"
RE_GROUP = re.compile(r"^\s*--\s*Group\s+(\d+)\s*--(?:\s+.*)?$", re.IGNORECASE)


def _style_header_row(ws, row_idx: int, bold=True, center=True):
    """Apply a consistent header style to a single row."""
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = WHITE_FONT if bold else Font(bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER if center else LEFT
        c.border = THIN_BORDER


def _style_group_header(ws, row_idx: int):
    """Style a group header row (the '-- Group N -- ...' row)."""
    max_col = max(1, ws.max_column)
    for col in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col)
        if col == 1:
            c.alignment = LEFT
            c.font = WHITE_FONT
        else:
            c.value = None  # keep single banner in col A
        c.fill = GROUP_FILL
        c.border = THIN_BORDER


def _autosize_columns(ws, min_width=10, padding=2):
    """Auto-fit column widths to contents with a minimum width."""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            ln = len(str(v))
            if ln + padding > max_len:
                max_len = ln + padding
        ws.column_dimensions[letter].width = max_len


# ----------------------------
# Key Sheet & Details Sheet
# ----------------------------

def add_key_sheet(wb, name, fields_desc: dict):
    sheet = wb.create_sheet(name)
    sheet.append(["Field", "Description", "Notes"])
    _style_header_row(sheet, 1)
    for field, val in fields_desc.items():
        if isinstance(val, tuple):
            if len(val) >= 2:
                desc, notes = val[0], val[1]
            else:
                desc, notes = val[0], ""
        else:
            desc = val
            notes = ""
        sheet.append([field, desc, notes])
    _autosize_columns(sheet)


def _append_details_sheet_if_present(wb, dest_name="Details"):
    details_path = os.environ.get("ROAD_DETAILS_XLSX", DETAILS_XLSX_DEFAULT)
    if not os.path.exists(details_path):
        return False, None
    try:
        src = openpyxl.load_workbook(details_path, data_only=False)
        if not src.sheetnames:
            return False, None
        src_ws = src[src.sheetnames[0]]
        name = dest_name
        base = dest_name
        i = 2
        while name in wb.sheetnames:
            name = f"{base} ({i})"; i += 1
        dst_ws = wb.create_sheet(name)
        for r in src_ws.iter_rows():
            for c in r:
                d = dst_ws.cell(row=c.row, column=c.col_idx, value=c.value)
                if c.has_style:
                    d.font = c.font.copy(); d.fill = c.fill.copy()
                    d.border = c.border.copy(); d.alignment = c.alignment.copy()
                    d.number_format = c.number_format; d.protection = c.protection.copy()
                if c.hyperlink: d.hyperlink = c.hyperlink.target or c.hyperlink
                if c.comment:   d.comment = openpyxl.comments.Comment(c.comment.text, c.comment.author or "")
        for m in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(m))
        for col, dim in src_ws.column_dimensions.items():
            if dim.width is not None:
                dst_ws.column_dimensions[col].width = dim.width
        for idx, dim in src_ws.row_dimensions.items():
            if dim.height is not None:
                dst_ws.row_dimensions[idx].height = dim.height
        dst_ws.freeze_panes = src_ws.freeze_panes
        if getattr(src_ws, "auto_filter", None) and src_ws.auto_filter.ref:
            dst_ws.auto_filter.ref = src_ws.auto_filter.ref
        return True, details_path
    except Exception:
        return False, None


# ----------------------------
# Export
# ----------------------------

def create_excel_from_bin(rengoku_data, output_file):
    """
    Export the parsed rengoku_data into a styled workbook:
      - Floor Stats (styled)
      - Spawn Table with group banners (styled) + optional group hints
      - Monster Key
      - Spawn Table Key
      - Optional Details sheet (copied from external xlsx)
    """
    spawn_tables, floor_stats, _, _, _, _ = rengoku_data
    wb = openpyxl.Workbook()

    # ---- Floor Stats
    ws_fs = wb.active
    ws_fs.title = "Floor Stats"
    fs_headers = ["FloorNumber", "SpawnTableUsed", "Unk0", "PointMulti1", "PointMulti2", "FinalLoop"]
    ws_fs.append(fs_headers)
    _style_header_row(ws_fs, 1)
    for stats in floor_stats:
        ws_fs.append([
            stats.FloorNumber,
            stats.SpawnTableUsed,
            stats.Unk0,
            stats.PointMulti1,
            stats.PointMulti2,
            stats.FinalLoop
        ])
    ws_fs.freeze_panes = "A2"
    ws_fs.auto_filter.ref = f"A1:{get_column_letter(len(fs_headers))}{ws_fs.max_row}"
    _autosize_columns(ws_fs)

    # ---- Spawn Tables
    ws_sp = wb.create_sheet("Spawn Table")
    sp_headers = [
        "FirstMonsterID", "FirstMonsterVariant",
        "SecondMonsterID", "SecondMonsterVariant",
        "MonstersStatTable", "Bonus Spawns",
        "SpawnWeighting", "AdditionalFlag"
    ]

    # Write groups with banners + per-group header rows
    for gi, group in enumerate(spawn_tables):
        # Banner row
        banner = f"-- Group {gi} --"
        hint = GROUP_HINTS.get(gi, "").strip()
        if hint:
            banner = f"{banner} {hint}"
        ws_sp.append([banner])
        _style_group_header(ws_sp, ws_sp.max_row)

        # Column header row (per group)
        ws_sp.append(sp_headers)
        _style_header_row(ws_sp, ws_sp.max_row)

        # Data rows
        for spawn in group:
            # spawn.output_excel_row(MONSTERS) must match sp_headers order
            ws_sp.append(spawn.output_excel_row(MONSTERS))

    # Optional: freeze top-left (won't follow each sub-header, but still helps)
    ws_sp.freeze_panes = "A2"
    _autosize_columns(ws_sp)

    # ---- How to Use
    how_to_use_fields = {
        "Overview": (
            "This workbook allows you to edit Hunting Road data extracted from `rengoku_data.bin`.",
            "It includes two main sheets — `Floor Stats` and `Spawn Table` — which define floor progression, monster combinations, and point multipliers."
        ),
        "Floor Stats Sheet": (
            "Contains one row per Hunting Road floor.",
            "Each row controls which spawn table is used and how many points are awarded. "
            "You can safely edit values here, but avoid deleting rows or renaming the sheet. "
            "View the Floor Stats Key for more information."
        ),
        "Spawn Table Sheet": (
            "Defines the monster combinations that can appear on each floor group.",
            "Each group begins with a header row, e.g., `-- Group 0 --`, followed by several monster entries. "
            "Do not delete or rename these group headers. "
            "Only edit the cells for Monster IDs, Variants, Stat Tables, and Flags."
        ),
        "Monster Key Reference": (
            "When editing `FirstMonsterID` or `SecondMonsterID`, enter the **monster name string**, not the numeric ID.",
            "You can find valid monster names in the `Monster Key` sheet — each corresponds to the internal EM ID used by the game. "
            "Example: instead of entering `11`, type `Rathalos`."
        ),
        "Editing Guidelines": (
            "You can edit cell values directly in Excel or Google Sheets.",
            "Make sure numerical fields (weights, multipliers, flags) contain valid numbers. "
            "Avoid leaving cells blank in important columns such as monster names or point multipliers."
        ),
        "Saving and Importing": (
            "After editing, save your file as `.xlsx` before importing.",
            "Do not change sheet names or remove tabs. "
            "When you import the edited file in the Hunting Road Editor, it will validate your data and write updates into a new BIN file."
        ),
        "Best Practices": (
            "Always keep a backup of your Excel before importing changes.",
            "Make small edits, save, and test them in-game after each change to ensure stability."
        ),
    }
    add_key_sheet(wb, "How to Use", how_to_use_fields)

    # ---- Monster Key
    ws_mk = wb.create_sheet("Monster Key")
    ws_mk.append(["EM ID", "Monster Name"])
    _style_header_row(ws_mk, 1)
    for i, monster in enumerate(MONSTERS):
        ws_mk.append([i, monster])
    ws_mk.freeze_panes = "A2"
    _autosize_columns(ws_mk)

    # ---- Floor Stats Key
    floor_stats_fields = {
        "FloorNumber": ("The sequential index of the floor within the Hunting Road.",),
        "SpawnTableUsed": ("The spawn table (Group) used for this floor.",),
        "Unk0": ("Unknown field — currently unused or reserved.",),
        "PointMulti1": (
            "Points multiplier applied to FirstMonsterID each floor, multiplies the internal base points for the monster.",),
        "PointMulti2": (
            "Points multiplier applied to SecondMonsterID each floor, multiplies the internal base points for the monster.",),
        "FinalLoop": ("Flag or loop index indicating when the final point loop begins (Typically floor 40).",)
    }
    add_key_sheet(wb, "Floor Stats Key", floor_stats_fields)
    # ---- Field Key
    spawn_table_fields = {
        "FirstMonsterID": ("ID of the first monster in the spawn pair",),
        "FirstMonsterVariant": ("Variant index of the first monster", "Different skins/forms or subspecies"),
        "SecondMonsterID": ("ID of the second monster in the spawn pair",),
        "SecondMonsterVariant": ("Variant index of the second monster", "Same as above but for the second monster"),
        "MonstersStatTable": ("Pointer/index to the monster stat table used", "Links to stats like HP, attack modifiers"),
        "Bonus Spawns": ("Stage-specific / bonus spawn parameters.",
                         "Default = 4294967295 \n0 = Shakalaka \n1 = Blango Spawns \n2 = King Shakalaka \n3 = Custom Spawn 1 \n4 = Custom Spawn 2 \n5 = Custom Spawn 3\n Note: Custom Spawn entries must be added manually into the questfile for Hunting Road, or using those parameters will crash."),
        "SpawnWeighting": ("Relative chance of this spawn being chosen", "Higher number = more likely"),
        "AdditionalFlag": ("Extra flags",
                           "0 = Default \n2 = Forced Spawn \n4 = Bonus Stage Flag \n6 = Forced Bonus Stage \n8 = Spawn Disabled ? (Needs Tests)"),
    }
    add_key_sheet(wb, "Spawn Table Key", spawn_table_fields)


    # ---- Optional Details sheet
    _append_details_sheet_if_present(wb, dest_name="Details")
    wb.save(output_file)


# ----------------------------
# Import
# ----------------------------

def export_excel_to_bin(excel_file, output_file, template_file):
    """
    Read a workbook that was exported by create_excel_from_bin()
    and write changes back into a new BIN, using the template_file as base.
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # ---- Parse Spawn Table sheet (robust group header detection)
    sp_ws = wb["Spawn Table"]
    tables = []
    spawn_group = []
    expecting_headers = False
    headers = []

    for row in sp_ws.iter_rows(values_only=True):
        first = row[0]

        # Detect group banner regardless of extra text after "-- Group N --"
        if isinstance(first, str) and RE_GROUP.match(first.strip()):
            # close previous group
            if spawn_group:
                tables.append(spawn_group)
                spawn_group = []
            expecting_headers = True
            headers = []
            continue

        # The row after a banner is a header row
        if expecting_headers:
            # Skip blank lines between banner & header (if user inserted any)
            if first is None or (isinstance(first, str) and not first.strip()):
                continue
            headers = [cell for cell in row]
            expecting_headers = False
            continue

        # Regular data rows (skip empty)
        if headers:
            if all(cell is None for cell in row):
                continue
            # Build dict for the row using headers
            item = {k: v for k, v in zip(headers, row)}
            spawn_group.append(item)

    if spawn_group:
        tables.append(spawn_group)

    # ---- Parse Floor Stats
    fs_ws = wb["Floor Stats"]
    stats = []
    headers = []
    for i, row in enumerate(fs_ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [cell for cell in row]
            continue
        if all(cell is None for cell in row):
            continue
        stat = {k: v for k, v in zip(headers, row)}
        stats.append(stat)

    # ---- Load template & apply changes
    from .io import parse_rengoku_data
    structs = parse_rengoku_data(template_file)
    with open(template_file, "rb") as f:
        data = bytearray(f.read())

    # Update spawn tables
    rengoku_tables = structs[0]
    # Zip will stop at the shortest; this prevents index errors if user adds/removes groups
    for table, group in zip(rengoku_tables, tables):
        for i, spawn in enumerate(table):
            if i >= len(group):
                break
            # Let the Spawn object read values from the row dict (names must match the exported headers)
            spawn.reset_values_from_row(MONSTERS, group[i])
            data[spawn.offset:spawn.offset + 32] = spawn.serialize()

    # Update floor stats
    floor_stats = structs[1]
    for i, stat in enumerate(stats):
        if i >= len(floor_stats):
            break
        fs = floor_stats[i]
        try:
            fs.FloorNumber   = int(stat.get("FloorNumber", fs.FloorNumber))
            fs.SpawnTableUsed= int(stat.get("SpawnTableUsed", fs.SpawnTableUsed))
            fs.Unk0          = int(stat.get("Unk0", fs.Unk0))
            fs.PointMulti1   = float(stat.get("PointMulti1", fs.PointMulti1))
            fs.PointMulti2   = float(stat.get("PointMulti2", fs.PointMulti2))
            fs.FinalLoop     = int(stat.get("FinalLoop", fs.FinalLoop))
        except Exception:
            # If any conversion fails, keep original values for that row
            pass
        data[fs.offset:fs.offset + 24] = fs.serialize()

    # ---- Write out the new BIN
    with open(output_file, "wb") as out_f:
        out_f.write(data)
