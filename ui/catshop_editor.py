# ui/catshop_editor.py
from __future__ import annotations

from pathlib import Path
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, Signal, QSortFilterProxyModel
from PySide6.QtGui import QPixmap, QPainter, QPalette, QBrush, QColor, QFont
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QGroupBox, QFormLayout, QSpinBox,
    QTableView, QHeaderView, QAbstractItemView, QHBoxLayout,
    QPushButton, QFileDialog, QMessageBox, QLineEdit, QGraphicsDropShadowEffect
)

from core.paths import ROOTDIR, resource_path
from ui.utils import apply_dialog_background
from core.catshop_io import parse_catshop, save_catshop, CatShopItem, CatShopParsed
from core.json_io import catshop_to_json, catshop_from_json
from .models import IntDelegate
from .models import EDITOR_TEXT_STYLE

# ---------- helpers ----------
def load_item_names() -> dict[int, str]:
    """
    Load item ID -> name from asset/Items.xlsx (first sheet).
    Expected headers include: ID (or ItemID), Name (or ItemName).
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    xlsx_path = resource_path("asset", "Items.xlsx")
    p = Path(xlsx_path)
    if not p.exists():
        p = Path("asset/Items.xlsx")
        if not p.exists():
            return {}

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = { (c.value or "").strip().lower(): idx
                    for idx, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))[0:]) }
        id_idx = None
        name_idx = None
        for key, idx in headers.items():
            if key in ("id", "itemid", "item_id"):
                id_idx = idx
            if key in ("name", "itemname", "item_name"):
                name_idx = idx
        if id_idx is None or name_idx is None:
            id_idx, name_idx = 0, 1

        mapping: dict[int, str] = {}
        for row in ws.iter_rows(min_row=2):
            try:
                rid = row[id_idx].value
                rname = row[name_idx].value
                if rid is None or rname is None:
                    continue
                mapping[int(rid)] = str(rname)
            except Exception:
                continue
        return mapping
    except Exception:
        return {}

class ItemListModel(QAbstractTableModel):
    COLS = ["id", "name"]
    HEADERS = ["ID", "Name"]

    def __init__(self, id_to_name: dict[int, str], parent=None):
        super().__init__(parent)
        self.rows = sorted([(i, n) for i, n in id_to_name.items()], key=lambda x: x[0])

    def rowCount(self, _=QModelIndex()):
        return len(self.rows)
    def columnCount(self, _=QModelIndex()):
        return 2

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self.HEADERS[section]
        return section + 1

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role in (Qt.DisplayRole, Qt.EditRole):
            rid, name = self.rows[index.row()]
            return rid if index.column() == 0 else name
        return None

class ItemListDialog(QDialog):
    def __init__(self, id_to_name: dict[int, str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("All Items (ID ↔ Name)")
        self.resize(520, 600)
        try:
            apply_dialog_background(self, image_name="bg2.jpg", opacity=0.65)
        except Exception:
            pass

        v = QVBoxLayout(self)
        label = QLabel("Search by ID or name", self)
        label.setAlignment(Qt.AlignCenter)
        v.addWidget(label)

        self.search = QLineEdit(self)
        self.search.setPlaceholderText("Type to filter…")
        v.addWidget(self.search)

        self.table = QTableView(self)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setDefaultSectionSize(22)
        self.table.setWordWrap(False)
        self.table.setShowGrid(True)
        v.addWidget(self.table, 1)

        self.model = ItemListModel(id_to_name, self)
        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.proxy.setSortCaseSensitivity(Qt.CaseInsensitive)
        self.proxy.setFilterKeyColumn(-1)
        self.table.setModel(self.proxy)

        hv = self.table.horizontalHeader()
        hv.setSectionResizeMode(QHeaderView.ResizeToContents)
        hv.setStretchLastSection(True)
        hv.setDefaultAlignment(Qt.AlignCenter)

        self.search.textChanged.connect(self._apply_filter)

    def _apply_filter(self, text: str):
        self.proxy.setFilterFixedString(text)

class CatShopModel(QAbstractTableModel):
    """
    Exposes: Item ID | Item Name | Item ID 2 | Item Name 2
    Only the ID columns are editable. Names come from a read-only mapping.
    """
    COLS = ["item_id", "item_name", "item_id2", "item_name2"]
    HEADERS = ["Item ID", "Item Name", "Item ID 2", "Item Name 2"]

    idsChanged = Signal()

    def __init__(self, rows: list[CatShopItem], id_to_name: dict[int, str], parent=None):
        super().__init__(parent)
        self.rows = rows
        self.id_to_name = id_to_name

    def rowCount(self, _=QModelIndex()) -> int:
        return len(self.rows)
    def columnCount(self, _=QModelIndex()) -> int:
        return len(self.COLS)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return self.HEADERS[section]
        return section + 1

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        col_name = self.COLS[index.column()]
        editable = col_name in ("item_id", "item_id2")
        base = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        return base | (Qt.ItemIsEditable if editable else Qt.NoItemFlags)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row = self.rows[index.row()]
        col_name = self.COLS[index.column()]

        if role in (Qt.DisplayRole, Qt.EditRole):
            if col_name == "item_id":
                return row.item_id
            elif col_name == "item_id2":
                return row.item_id2
            elif col_name == "item_name":
                return self.id_to_name.get(int(row.item_id), f"Unknown ({row.item_id})")
            elif col_name == "item_name2":
                return self.id_to_name.get(int(row.item_id2), f"Unknown ({row.item_id2})")

        if role == Qt.ForegroundRole:
            if col_name in ("item_id", "item_id2"):
                return QColor("#66CCFF")
            elif col_name in ("item_name", "item_name2"):
                return QColor("#FFFFFF")

        if role == Qt.FontRole:
            if col_name in ("item_id", "item_id2"):
                font = QFont()
                font.setBold(True)
                return font
        return None

    def setData(self, index, value, role=Qt.EditRole):
        if role != Qt.EditRole or not index.isValid():
            return False
        row = self.rows[index.row()]
        col_name = self.COLS[index.column()]
        try:
            ival = int(value)
            changed = False
            if col_name == "item_id":
                if row.item_id != ival:
                    row.item_id = ival
                    changed = True
            elif col_name == "item_id2":
                if row.item_id2 != ival:
                    row.item_id2 = ival
                    changed = True
            else:
                return False
            if changed:
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                self.idsChanged.emit()
            return True
        except Exception:
            return False

    def begin_full_reset(self):
        self.beginResetModel()
    def end_full_reset(self):
        self.endResetModel()

class CatShopEditor(QDialog):
    """
    Road Cat Item Shop editor:
    """
    def __init__(self, mhfdat_path: str, mhfdat_parsed: dict, parent=None):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowTitle("Road Cat Item Shop")
        self.resize(1000, 640)

        try:
            apply_dialog_background(self, image_name="bg2.jpg", opacity=0.65)
        except Exception:
            bg = QPixmap(str(ROOTDIR / "asset/bg2.jpg"))
            if not bg.isNull():
                bg = bg.scaled(self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                tmp = QPixmap(bg.size()); tmp.fill(Qt.transparent)
                p = QPainter(tmp); p.setOpacity(0.65); p.drawPixmap(0, 0, bg); p.end()
                pal = self.palette(); pal.setBrush(QPalette.ColorRole.Window, QBrush(tmp)); self.setPalette(pal)

        self.setWindowFlags(
            Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint |
            Qt.WindowMaximizeButtonHint | Qt.WindowTitleHint
        )

        self.mhfdat_path = mhfdat_path
        self.mhfdat_parsed = mhfdat_parsed or {}
        self.counters = self.mhfdat_parsed.get("counters")

        self.id_to_name = load_item_names()

        parsed = parse_catshop(mhfdat_path) or CatShopParsed(rows=[])
        self.parsed = parsed

        root = QVBoxLayout(self)

        header = QLabel("Road Cat Item Shop", self)
        header.setStyleSheet("Color: cyan;")
        header.setAlignment(Qt.AlignCenter)
        header.setProperty("class", "header")
        root.addWidget(header)

        glow = QGraphicsDropShadowEffect(self)
        glow.setBlurRadius(20)
        glow.setColor(QColor("#00FFFF"))
        glow.setOffset(0, 0)
        header.setGraphicsEffect(glow)

        grp = QGroupBox("Shop Counter", self)
        form = QFormLayout(grp)
        grp.setStyleSheet("color: cyan;")
        label = QLabel("Total Items:", self)
        label.setStyleSheet("color: #66CCFF")
        self.lbl_count = QLabel("0", grp)
        self.lbl_count.setStyleSheet("color: cyan; font-weight: bold;")
        self.lbl_count.setToolTip("Auto-syncs to the total number of items (Item ID + Item ID 2 across all rows).")
        form.addRow(label, self.lbl_count)
        root.addWidget(grp)

        self.table = QTableView(self)
        self._style_table(self.table)
        root.addWidget(self.table, 1)

        self.model = CatShopModel(self.parsed.rows, self.id_to_name, self)
        self.model.idsChanged.connect(self._update_counter_from_rows)
        self.table.setModel(self.model)
        self.table.setItemDelegateForColumn(0, IntDelegate(0, 65535, self.table))
        self.table.setItemDelegateForColumn(2, IntDelegate(0, 65535, self.table))

        hv = self.table.horizontalHeader()
        hv.setSectionResizeMode(QHeaderView.ResizeToContents)
        hv.setStretchLastSection(True)
        hv.setDefaultAlignment(Qt.AlignLeft)
        row = QHBoxLayout()
        row.addStretch(1)

        self.btn_add = QPushButton("Add Row", self)
        self.btn_add.clicked.connect(self._add_row)
        row.addWidget(self.btn_add)

        self.btn_del = QPushButton("Delete Selected", self)
        self.btn_del.clicked.connect(self._delete_selected)
        row.addWidget(self.btn_del)

        self.btn_items_list = QPushButton("Items List", self)
        self.btn_items_list.clicked.connect(self._show_items_dialog)
        row.addWidget(self.btn_items_list)

        self.btn_export_json = QPushButton("Export JSON", self)
        self.btn_export_json.clicked.connect(self._export_json)
        row.addWidget(self.btn_export_json)

        self.btn_import_json = QPushButton("Import JSON", self)
        self.btn_import_json.clicked.connect(self._import_json)
        row.addWidget(self.btn_import_json)

        self.btn_save = QPushButton("Save", self)
        self.btn_save.clicked.connect(self._save)
        row.addWidget(self.btn_save)

        self.setStyleSheet("""
            QPushButton {
                font-weight: bold;
            }
        """)
        row.addStretch(1)
        root.addLayout(row)

        self._update_counter_from_rows()

    def _style_table(self, tv: QTableView):
        tv.setAlternatingRowColors(True)
        tv.setSortingEnabled(False)
        tv.setSelectionBehavior(QAbstractItemView.SelectRows)
        tv.setSelectionMode(QAbstractItemView.SingleSelection)
        tv.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.SelectedClicked
            | QAbstractItemView.EditKeyPressed
        )
        tv.verticalHeader().setVisible(True)
        tv.verticalHeader().setDefaultSectionSize(24)
        tv.setWordWrap(False)
        tv.setShowGrid(True)

    def _refresh_model(self):
        self.model.begin_full_reset()
        self.model.end_full_reset()
        self._update_counter_from_rows()

    def _computed_item_count(self) -> int:
        total = 0
        for r in self.parsed.rows:
            if int(r.item_id) != 0:
                total += 1
            if int(r.item_id2) != 0:
                total += 1
        return total

    def _update_counter_from_rows(self):
        total = self._computed_item_count()
        self.lbl_count.setText(str(total))
        if self.counters:
            self.counters.c_unk3 = total

    def _add_row(self):
        for row in self.parsed.rows:
            if int(row.item_id2) == 0:
                QMessageBox.warning(
                    self, "Invalid Add",
                    "All rows must have both Item ID and Item ID 2 filled before adding another row."
                )
                return
        self.parsed.rows.append(CatShopItem(item_id=0, item_id2=0))
        self._refresh_model()

    def _delete_selected(self):
        idx = self.table.currentIndex()
        if not idx.isValid():
            return
        r = idx.row()
        if 0 <= r < len(self.parsed.rows):
            del self.parsed.rows[r]
            self._refresh_model()

    def _show_items_dialog(self):
        if not self.id_to_name:
            QMessageBox.information(self, "No item mapping",
                                    "No item names were loaded from asset/Items.xlsx.")
            return
        dlg = ItemListDialog(self.id_to_name, self)
        dlg.exec()

    def _export_json(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export CatShop JSON", "catshop.json", "JSON Files (*.json)")
        if not path:
            return
        try:
            s = catshop_to_json(self.parsed)
            with open(path, "w", encoding="utf-8") as f:
                f.write(s)
            QMessageBox.information(self, "Exported", f"Exported CatShop JSON to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export JSON:\n{e}")

    def _import_json(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import CatShop JSON", "", "JSON Files (*.json)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                s = f.read()
            new_parsed = catshop_from_json(s)
            self.parsed = new_parsed
            self.model.rows = self.parsed.rows
            self.model.begin_full_reset()
            self.model.end_full_reset()
            self._update_counter_from_rows()
            QMessageBox.information(self, "Imported", f"Imported CatShop JSON from {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import JSON:\n{e}")

    def _save(self):
        for i, row in enumerate(self.parsed.rows):
            if i < len(self.parsed.rows) - 1:
                if int(row.item_id) == 0 or int(row.item_id2) == 0:
                    QMessageBox.warning(
                        self, "Invalid Save",
                        f"Row {i + 1} has an empty Item ID. Please fill it before saving."
                    )
                    return

        out_path, _ = QFileDialog.getSaveFileName(
            self, "Save mhfdat.bin", "mhfdat.bin", "Binary Files (*.bin)"
        )
        if not out_path:
            return

        try:
            computed = self._computed_item_count()
            save_catshop(
                self.mhfdat_path,
                out_path,
                self.parsed,
                counters=self.counters,
                counter_items_count=computed,
                always_move_to_eof=True,
                eof_align=0x10,
                end_padding=0x400
            )
            QMessageBox.information(
                self, "Saved",
                "Cat Shop written to EOF, pointer (0xB10) updated, and counter synced."
            )
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", f"Failed to save Cat Shop:\n{e}")
