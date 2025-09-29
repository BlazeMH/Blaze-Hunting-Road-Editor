import sys
import os
import struct
import re
import openpyxl

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

PROJECTDIR = os.path.dirname(__file__)


MONSTERS = [
    "None",
    "Rathian",
    "Fatalis",
    "Kelbi",
    "Mosswine",
    "Bullfango",
    "Yian_Kut_Ku",
    "Lao_Shan_Lung",
    "Cephadrome",
    "Felyne_1",
    "Veggie_Elder",
    "Rathalos",
    "Aptonoth",
    "Genprey",
    "Diablos",
    "Khezu",
    "Velociprey",
    "Gravios",
    "Felyne_2",
    "Vespoid",
    "Gypceros",
    "Plesioth",
    "Basarios",
    "Melynx",
    "Hornetaur",
    "Apceros",
    "Monoblos",
    "Velocidrome",
    "Gendrome",
    "Rocks_0",
    "Ioprey",
    "Iodrome",
    "Pugis",
    "Kirin",
    "Cephalos",
    "Giaprey",
    "Crimson_Fatalis",
    "Pink_Rathian",
    "Blue_Yian_Kut_Ku",
    "Purple_Gypceros",
    "Yian_Garuga",
    "Silver_Rathalos",
    "Gold_Rathian",
    "Black_Diablos",
    "White_Monoblos",
    "Red_Khezu",
    "Green_Plesioth",
    "Black_Gravios",
    "Daimyo_Hermitaur",
    "Azure_Rathalos",
    "Ashen_Lao_Shan_Lung",
    "Blangonga",
    "Congalala",
    "Rajang",
    "Kushala_Daora",
    "Shen_Gaoren",
    "Great_Thunderbug",
    "Shakalaka",
    "Yama_Tsukami_1",
    "Chameleos",
    "Rusted_Kushala_Daora",
    "Blango",
    "Conga",
    "Remobra",
    "Lunastra",
    "Teostra",
    "Hermitaur",
    "Shogun_Ceanataur",
    "Bulldrome",
    "Anteka",
    "Popo",
    "White_Fatalis",
    "Yama_Tsukami_2",
    "Ceanataur",
    "Hypnocatrice",
    "Lavasioth",
    "Tigrex",
    "Akantor",
    "Bright_Hypnoc",
    "Lavasioth_Subspecies",
    "Espinas",
    "Orange_Espinas",
    "White_Hypnoc",
    "Akura_Vashimu",
    "Akura_Jebia",
    "Berukyurosu",
    "Cactus_01",
    "Gorge_Objects",
    "Gorge_Rocks",
    "Pariapuria",
    "White_Espinas",
    "Kamu_Orugaron",
    "Nono_Orugaron",
    "Raviente",
    "Dyuragaua",
    "Doragyurosu",
    "Gurenzeburu",
    "Burukku",
    "Erupe",
    "Rukodiora",
    "Unknown",
    "Gogomoa",
    "Kokomoa",
    "Taikun_Zamuza",
    "Abiorugu",
    "Kuarusepusu",
    "Odibatorasu",
    "Disufiroa",
    "Rebidiora",
    "Anorupatisu",
    "Hyujikiki",
    "Midogaron",
    "Giaorugu",
    "Mi_Ru",
    "Farunokku",
    "Pokaradon",
    "Shantien",
    "Pokara",
    "Dummy",
    "Goruganosu",
    "Aruganosu",
    "Baruragaru",
    "Zerureusu",
    "Gougarf",
    "Uruki",
    "Forokururu",
    "Meraginasu",
    "Diorekkusu",
    "Garuba_Daora",
    "Inagami",
    "Varusaburosu",
    "Poborubarumu",
    "Duremudira",
    "UNK_0",
    "Felyne",
    "Blue_NPC",
    "UNK_1",
    "Cactus_Varusa",
    "Veggie_Elders",
    "Gureadomosu",
    "Harudomerugu",
    "Toridcless",
    "Gasurabazura",
    "Kusubami",
    "Yama_Kurai",
    "Dure_2nd_District",
    "Zinogre",
    "Deviljho",
    "Brachydios",
    "Berserk_Laviente",
    "Toa_Tesukatora",
    "Barioth",
    "Uragaan",
    "Stygian_Zinogre",
    "Guanzorumu",
    "Starving_Deviljho",
    "UNK",
    "Egyurasu",
    "Voljang",
    "Nargacuga",
    "Keoaruboru",
    "Zenaserisu",
    "Gore_Magala",
    "Blinking_Nargacuga",
    "Shagaru_Magala",
    "Amatsu",
    "Elzelion",
    "Musou_Dure",
    "Rocks_1",
    "Seregios",
    "Bogabadorumu",
    "Unknown_Blue_Barrel",
    "Musou_Bogabadorumu",
    "Costumed_Uruki",
    "Musou_Zerureusu",
    "PSO2_Rappy",
    "King_Shakalaka"
]

def add_key_sheet(wb, name, fields_desc: dict):
    sheet = wb.create_sheet(name)
    sheet.append(["Field", "Description", "Notes"])
    for field, val in fields_desc.items():
        if isinstance(val, tuple):
            # expect (description, notes)
            if len(val) >= 2:
                desc, notes = val[0], val[1]
            else:
                desc, notes = val[0], ""
        else:
            desc = val
            notes = ""
        sheet.append([field, desc, notes])

class RoadMode:

    def __init__(self, data, offset):
        (self.FloorStatsCount, self.SpawnCountCount, self.SpawnTablePointersCount,
        self.FloorStatsPointer, self.SpawnTablePointers, self.SpawnCountPointers) = struct.unpack('<6I', data)
        self.offset = offset

    def addSpawnTables(self, obj):
        self.spawnTables = obj

    def addFloorStats(self, obj):
        self.floorStats = obj

    def todict(self):
        return {
            'FloorStatsCount': self.FloorStatsCount,
            'SpawnCountCount': self.SpawnCountCount,
            'SpawnTablePointersCount': self.SpawnTablePointersCount,
            'FloorStatsPointer': self.FloorStatsPointer,
            'SpawnTablePointers': self.SpawnTablePointers,
            'SpawnCountPointers': self.SpawnCountPointers
        }

    def serialize(self):
        return struct.pack('<6I', self.FloorStatsCount, self.SpawnCountCount, self.SpawnTablePointersCount,
        self.FloorStatsPointer, self.SpawnTablePointers, self.SpawnCountPointers)



    def __str__(self):
        return str(self.todict())

    __repr__ = __str__


class SpawnTable:

    title = "SpawnTable"
    def __init__(self, data, offset):
        self.FirstMonsterID, self.FirstMonsterVariant, self.SecondMonsterID, self.SecondMonsterVariant, \
        self.MonstersStatTable, self.MapZoneOverride, self.SpawnWeighting, self.AdditionalFlag = struct.unpack('<2I2I4I', data)
        self.offset = offset

    def todict(self):
        return {
            'FirstMonsterID': self.FirstMonsterID,
            'FirstMonsterVariant': self.FirstMonsterVariant,
            'SecondMonsterID': self.SecondMonsterID,
            'SecondMonsterVariant': self.SecondMonsterVariant,
            'MonstersStatTable': self.MonstersStatTable,
            'MapZoneOverride': self.MapZoneOverride,
            'SpawnWeighting': self.SpawnWeighting,
            'AdditionalFlag': self.AdditionalFlag
        }

    def __str__(self):
        return str(self.todict())

    def output_excel_row(self):
        return [
            MONSTERS[self.FirstMonsterID],
            self.FirstMonsterVariant,
            MONSTERS[self.SecondMonsterID],
            self.SecondMonsterVariant,
            self.MonstersStatTable,
            self.MapZoneOverride,
            self.SpawnWeighting,
            self.AdditionalFlag
        ]

    def check_monster_id(self, monster_id):
        if isinstance(monster_id, str):
            if monster_id.isdigit():
                monster_id = int(monster_id)
            else:
                monster_id = MONSTERS.index(monster_id)
        return monster_id

    def reset_values_from_row(self, group):
        self.FirstMonsterID = self.check_monster_id(group["FirstMonsterID"])
        self.FirstMonsterVariant = int(group["FirstMonsterVariant"])
        self.SecondMonsterID = self.check_monster_id(group["SecondMonsterID"])
        self.SecondMonsterVariant = int(group["SecondMonsterVariant"])
        self.MonstersStatTable = int(group["MonstersStatTable"])
        self.MapZoneOverride = int(group["MapZoneOverride"])
        self.SpawnWeighting = int(group["SpawnWeighting"])
        self.AdditionalFlag = int(group["AdditionalFlag"])


    def serialize(self):
        return struct.pack('<2I2I4I', self.FirstMonsterID, self.FirstMonsterVariant, self.SecondMonsterID, self.SecondMonsterVariant,
        self.MonstersStatTable, self.MapZoneOverride, self.SpawnWeighting, self.AdditionalFlag)

    __repr__ = __str__


class FloorStats:
    title = "FloorStats"
    def __init__(self, data, offset):
        self.FloorNumber, self.SpawnTableUsed, self.Unk0, self.PointMulti1, self.PointMulti2, self.FinalLoop = struct.unpack('<3I2fI', data)
        self.offset = offset

    def todict(self):
        return {
            'FloorNumber': self.FloorNumber,
            'SpawnTableUsed': self.SpawnTableUsed,
            'Unk0': self.Unk0,
            'PointMulti1': self.PointMulti1,
            'PointMulti2': self.PointMulti2,
            'FinalLoop': self.FinalLoop
        }

    def serialize(self):
        return struct.pack('<3I2fI', self.FloorNumber, self.SpawnTableUsed, self.Unk0, self.PointMulti1, self.PointMulti2, self.FinalLoop)

    def reset_values_from_row(self, row):
        self.FloorNumber = int(row["FloorNumber"])
        self.SpawnTableUsed = int(row["SpawnTableUsed"])
        self.Unk0 = int(row["Unk0"])
        self.PointMulti1 = float(row["PointMulti1"])
        self.PointMulti2 = float(row["PointMulti2"])
        self.FinalLoop = int(row["FinalLoop"])

    def __str__(self):
        return str(self.todict())

    __repr__ = __str__


def parse_rengoku_data(file_path):
    if not os.path.exists(file_path):
        return
    with open(file_path, 'rb') as f:
        f.seek(0x14)
        offset = f.tell()
        multi_def = RoadMode(f.read(24), offset)
        offset = f.tell()
        solo_def = RoadMode(f.read(24), offset)

        # Parsing multiDef spawn tables
        spawn_tables = []
        for i in range(multi_def.SpawnTablePointersCount):
            f.seek(multi_def.SpawnTablePointers + (i * 4))
            table_pointer = struct.unpack('<I', f.read(4))[0]
            f.seek(multi_def.SpawnCountPointers + (i * 4))
            entry_count = struct.unpack('<I', f.read(4))[0]
            spawns = []
            f.seek(table_pointer)
            for x in range(entry_count):
                offset = f.tell()
                spawns.append(SpawnTable(f.read(32), offset))
            spawn_tables.append(spawns)
        multi_def.addSpawnTables(spawn_tables)

        # Parsing multiDef floor stats
        floor_stats = []
        f.seek(multi_def.FloorStatsPointer)
        for i in range(multi_def.FloorStatsCount):
            offset = f.tell()
            floor_stats.append(FloorStats(f.read(24), offset))
        multi_def.addFloorStats(floor_stats)

        # Parsing soloDef spawn tables
        spawn_tables_solo = []
        for i in range(solo_def.SpawnTablePointersCount):
            f.seek(solo_def.SpawnTablePointers + (i * 4))
            table_pointer = struct.unpack('<I', f.read(4))[0]
            f.seek(solo_def.SpawnCountPointers + (i * 4))
            entry_count = struct.unpack('<I', f.read(4))[0]
            spawns = []
            f.seek(table_pointer)
            for x in range(entry_count):
                offset = f.tell()
                spawns.append(SpawnTable(f.read(32), offset))
            spawn_tables_solo.append(spawns)
        solo_def.addSpawnTables(spawn_tables_solo)
        # Parsing soloDef floor stats
        floor_stats_solo = []
        f.seek(solo_def.FloorStatsPointer)
        for i in range(solo_def.FloorStatsCount):
            offset = f.tell()
            floor_stats_solo.append(FloorStats(f.read(24), offset))
        solo_def.addFloorStats(floor_stats_solo)
    return [spawn_tables, floor_stats, multi_def, spawn_tables_solo, floor_stats_solo, solo_def]


def create_excel_from_bin(rengoku_data, output_file):
    spawn_tables, floor_stats, multi_def, spawn_tables_solo, floor_stats_solo, solo_def = rengoku_data
    wb = openpyxl.Workbook()

    wb.active.title = "Floor Stats"
    floor_stats_sheet = wb.active
    headers = ["FloorNumber", "SpawnTableUsed", "Unk0", "PointMulti1", "PointMulti2", "FinalLoop"]
    floor_stats_sheet.append(headers)

    for stats in floor_stats:
        floor_stats_sheet.append([stats.FloorNumber, stats.SpawnTableUsed, stats.Unk0,
                                  stats.PointMulti1, stats.PointMulti2, stats.FinalLoop])


    spawn_table_sheet = wb.create_sheet("Spawn Table")
    headers = ["FirstMonsterID", "FirstMonsterVariant", "SecondMonsterID", "SecondMonsterVariant",
            "MonstersStatTable", "MapZoneOverride", "SpawnWeighting", "AdditionalFlag"]

    for group in spawn_tables:
        spawn_table_sheet.append(["-- Group {} --".format(spawn_tables.index(group))])
        spawn_table_sheet.append(headers)
        for spawn in group:
            spawn_table_sheet.append(spawn.output_excel_row())

    # Create Floor Stats sheet
    monster_key = wb.create_sheet("Monster Key")
    monster_key.append(["EM ID", "Monster Name"])
    for i, monster in enumerate(MONSTERS):
        monster_key.append([i, monster])



    # Create Details Key
    spawn_table_fields = {
        "FirstMonsterID": ("ID of the first monster in the spawn pair",),
        "FirstMonsterVariant": ("Variant index of the first monster", "Different skins/forms or subspecies"),
        "SecondMonsterID": ("ID of the second monster in the spawn pair",),
        "SecondMonsterVariant": ("Variant index of the second monster", "Same as above but for the second monster"),
        "MonstersStatTable": ("Pointer/index to the monster stat table used",
                              "Links to stats like HP, attack modifiers"),
        "MapZoneOverride": ("In the context of Hunting Road, this field refers to special parameters set for stages.","0 = Shakalaka \n1 = Blango Spawns \n2 = King Shakalaka \nNote: It is possible to add additional Bonus Floor Options by explicitly adding MapZoneOverride entries in the Hunting Road Quest File (23527d0.bin)"),
        "SpawnWeighting": ("Relative chance of this spawn being chosen", "Higher number = more likely"),
        "AdditionalFlag": (
            "Extra flags",
            """0 = Default \n2 = Forced Spawn (used on Fatalis Floors for example) \n4 = Bonus Stage Flag (This Flag enables Purple Text + Road Medal Award) \n6 = Forced Bonus Stage Spawn (This Flag forces the Spawn and Bonus Stage Parameters.) \n8 = Spawn Disabled ? (More Testing Needed)
            \nNote: Flags 1, 3, and 5 are undiscovered. If anyone has information on these, please share."""
        ),
    }
    add_key_sheet(wb, "Spawn Table Key", spawn_table_fields)

    # Save the workbook
    wb.save(output_file)


def export_excel_to_bin(excel_file, output_file, template_file):
    wb = openpyxl.load_workbook(excel_file)
    spawn_tables = wb["Spawn Table"]
    tables = []
    spawn_group = []
    headers_row = False
    headers = []
    group_num = 0
    for i, row in enumerate(spawn_tables.rows):
        if re.match(r"-- Group \d+ --", str(row[0].value)):
            headers_row = True
            group_num = int(row[0].value.split(" ")[2])
            if spawn_group:
                tables.append(spawn_group)
                spawn_group = []
        elif headers_row == True:
            headers = [cell.value for cell in row]
            headers_row = False
        else:
            spawn_table = {i:v for i, v in zip(headers, [cell.value for cell in row])}
            spawn_group.append(spawn_table)
    floor_stats = wb["Floor Stats"]
    stats = []
    for i, row in enumerate(floor_stats.rows):
        if i == 0:
            headers = [cell.value for cell in row]
        else:
            stat = {i:v for i,v in zip(headers, [cell.value for cell in row])}
            stats.append(stat)
    structs = parse_rengoku_data(template_file)
    with open(template_file, "rb") as f:
        data = f.read()
        data = bytearray(data)
    rengoku_tables = structs[0]
    for table, group in zip(rengoku_tables, tables):
        for i, spawn in enumerate(table):
            if i >= len(group):
                break
            spawn.reset_values_from_row(group[i])
            data[spawn.offset:spawn.offset + 32] = spawn.serialize()
    floor_stats = structs[1]
    for i, stat in enumerate(stats):
        if i >= len(floor_stats):
            break
        floor_stats[i].reset_values_from_row(stat)
        data[floor_stats[i].offset:floor_stats[i].offset + 24] = floor_stats[i].serialize()
    with open(output_file, "wb") as out_f:
            out_f.write(data)


style = """
QPushButton {
    background-color: #1c0c29;
    color: white;
    padding: 10px 20px;
    border: 1px solid #4CA;
    border-radius: 5px;
    font-size: 16px;
    width: 250px;
}
QPushButton:hover {
    background-color: #502275;
    color: yellow;
    padding: 10px 20px;
    border: 1px solid #134639;
    border-radius: 5px;
    font-size: 16px;
    width: 250px;
}
QPushButton:pressed {
    background-color: #7130a5;
    color: dark-yellow;
    padding: 10px 20px;
    border: 1px solid #0c2922;
    border-radius: 5px;
    font-size: 16px;
    width: 250px;
}
"""

class RengokuWindow(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setWindowTitle("Blaze Road Editor")
        self.setGeometry(100, 100, 800, 600)
        self.initUI()

    def initUI(self):
        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)
        self.setWindowIcon(QIcon(os.path.join(PROJECTDIR, "../asset/icon.png")))
        self.bakimg = QPixmap(os.path.join(PROJECTDIR, "../asset/bg.png"))
        self.bakimg = self.bakimg.scaled(self.size(), Qt.AspectRatioMode.KeepAspectRatioByExpanding, Qt.TransformationMode.SmoothTransformation)
        self.palette = QPalette()
        self.palette.setBrush(QPalette.ColorRole.Window, QBrush(self.bakimg))
        self.setPalette(self.palette)
        self.hlayout = QHBoxLayout(self.central_widget)
        self.layout = QVBoxLayout()
        self.hlayout.addStretch(1)
        self.hlayout.addLayout(self.layout)
        self.hlayout.addStretch(1)
        self.layout.addStretch(1)

        self.load_button = QPushButton("Load Rengoku Data", self)
        self.load_button.clicked.connect(self.load_rengoku_data)
        self.layout.addWidget(self.load_button)

        self.export_button = QPushButton("Export to Excel", self)
        self.export_button.clicked.connect(self.export_to_excel)
        self.layout.addWidget(self.export_button)

        self.import_button = QPushButton("Import from Excel", self)
        self.import_button.clicked.connect(self.import_from_excel)
        self.layout.addWidget(self.import_button)
        self.layout.addStretch(1)

    def load_rengoku_data(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Rengoku Data File", "", "Binary Files (*.bin)")
        if not file_path:
            return

        # decompression check: if the file is literally named rengoku_data.bin and is tiny, it's probably compressed
        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            QMessageBox.critical(self, "Error", "Failed to read the file size.")
            return

        if os.path.basename(file_path).lower() == 'rengoku_data.bin' and file_size < 10 * 1024:
            QMessageBox.critical(
                self,
                "Error",
                "The selected Rengoku file appears to be compressed or truncated (size < 10 KB). Please decompress it and try again."
            )
            return

        # proceed with parsing
        self.rengoku_path = file_path
        self.rengoku_data = parse_rengoku_data(file_path)
        if self.rengoku_data:
            QMessageBox.information(self, "Success", "Rengoku data loaded successfully!")
        else:
            QMessageBox.critical(self, "Error", "Failed to parse Rengoku data.")

    def export_to_excel(self):
        if hasattr(self, 'rengoku_data'):
            output_file, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
            if output_file:
                create_excel_from_bin(self.rengoku_data, output_file)
                QMessageBox.information(self, "Success", "Data exported to Excel successfully!")
        else:
            QMessageBox.warning(self, "Error", "No Rengoku data loaded!")

    def import_from_excel(self):
        excel_file, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        if excel_file:
            output_file, _ = QFileDialog.getSaveFileName(self, "Save Rengoku Data File", "", "Binary Files (*.bin)")
            if output_file:
                if hasattr(self, "rengoku_path"):
                    template_file = self.rengoku_path
                else:
                    template_file, _ = QFileDialog.getOpenFileName(self, "Open Rengoku Template File", "", "Binary Files (*.bin)")
                export_excel_to_bin(excel_file, output_file, template_file)
                QMessageBox.information(self, "Success", "Data imported from Excel successfully!")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(style)
    window = RengokuWindow()
    window.show()
    sys.exit(app.exec())
